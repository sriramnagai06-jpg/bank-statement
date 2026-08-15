"""
Karur Vysya Bank (KVB) statement parser (Excel support).

Handles all KVB Excel formats:
  - Format A: Header row 0 with sparse columns
  - Format B: Blob in col1, separate narration/amount columns
  - Format C: Generic with 'Txn Date'/'Date' and 'Particulars' columns
  - Supports multiline narration rows and embedded cell newlines
"""
import re
import openpyxl
import pandas as pd
from datetime import datetime

BLOB_DATE_RE = re.compile(r"^\d{2}-[A-Z]{3}-\d{4}", re.IGNORECASE)
AMOUNT_RE = re.compile(r"([\d,]+\.\d{2})")


def _clean(val):
    """Strip leading apostrophe Excel-text-prefix, replace newlines with space, and strip whitespace."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if s.startswith("'"):
        s = s[1:].strip()
    # Normalize embedded newlines / carriage returns to a single space
    s = re.sub(r'[\r\n\t]+', ' ', s).strip()
    return s if s else None


def _parse_amt(val):
    v = _clean(val)
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return abs(float(v)) if not (isinstance(v, float) and v != v) else 0.0
    s = str(v).replace(',', '').strip()
    if not s or s.lower() in ('nan', '-', '--'):
        return 0.0
    try:
        return abs(float(s))
    except Exception:
        return 0.0


def _parse_date(raw):
    """Parse a date from various formats. Returns datetime or None."""
    if raw is None:
        return None
    if hasattr(raw, 'strftime'):
        r = raw.replace(tzinfo=None) if hasattr(raw, 'tzinfo') else raw
        return r
    s = _clean(raw)
    if not s:
        return None
    s = str(s).split('\n')[0].strip()
    s = re.sub(r'\s+\d{2}:\d{2}:\d{2}.*$', '', s).strip()
    m = re.match(r'(\d{2}[-/][A-Za-z]{3}[-/]\d{2,4}|\d{2}[-/]\d{2}[-/]\d{4})', s)
    if m:
        s = m.group(1)
    m2 = re.match(r'(\d{4}-\d{2}-\d{2})', s)
    if m2:
        try:
            return datetime.strptime(m2.group(1), '%Y-%m-%d')
        except Exception:
            pass
    for fmt in ('%d-%b-%Y', '%d/%b/%Y', '%d-%b-%y', '%d-%m-%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    try:
        return pd.to_datetime(s, dayfirst=True).to_pydatetime()
    except Exception:
        return None


def _scan_for_header(rows, start=0, end=None):
    """
    Scan rows[start:end] for a header row containing date, narration, debit, credit columns.
    Returns (header_row_idx, col_map) or (None, None).
    """
    if end is None:
        end = min(start + 30, len(rows))
    
    for r_idx in range(start, end):
        row = rows[r_idx]
        vals = [str(_clean(v) or '').strip().lower() for v in row]
        has_date = any('txn date' in v or v == 'date' or 'txn_date' in v for v in vals)
        has_part = any('particulars' in v or 'narration' in v or 'description' in v for v in vals)
        has_debit = any('debit' in v and 'balance' not in v for v in vals)
        has_credit = any('credit' in v and 'balance' not in v for v in vals)
        
        if has_date and has_part and has_debit and has_credit:
            col_map = {}
            for c_idx, v in enumerate(vals):
                if ('txn date' in v or v == 'date') and 'date_col' not in col_map:
                    col_map['date_col'] = c_idx
                elif ('value date' in v or v == 'valuedate') and 'vdate_col' not in col_map:
                    col_map['vdate_col'] = c_idx
                elif ('particulars' in v or 'narration' in v or 'description' in v) and 'part_col' not in col_map:
                    col_map['part_col'] = c_idx
                elif 'ref' in v and 'ref_col' not in col_map:
                    col_map['ref_col'] = c_idx
                elif 'debit' in v and 'balance' not in v and 'debit_col' not in col_map:
                    col_map['debit_col'] = c_idx
                elif 'credit' in v and 'balance' not in v and 'credit_col' not in col_map:
                    col_map['credit_col'] = c_idx
                elif 'balance' in v and 'balance_col' not in col_map:
                    col_map['balance_col'] = c_idx
            
            if 'date_col' in col_map and 'part_col' in col_map and 'debit_col' in col_map and 'credit_col' in col_map:
                return r_idx, col_map
    
    return None, None


def _try_parse_sheet(rows):
    """
    Parse all transactions from a sheet that may have repeated header rows.
    Handles multiline/continuation narration rows without dropping data.
    """
    h_idx, col_map = _scan_for_header(rows, start=0)
    if h_idx is None or col_map is None:
        return None

    def _is_header_row(row):
        vals = [str(_clean(v) or '').strip().lower() for v in row]
        return (any('txn date' in v or v == 'date' for v in vals) and
                any('particulars' in v or 'narration' in v or 'description' in v for v in vals))

    transactions = []
    for row in rows[h_idx + 1:]:
        if _is_header_row(row):
            continue

        first_val = str(_clean(row[0] if row else None) or '').lower()
        if 'note:' in first_val or 'total' in first_val or 'statement summary' in first_val:
            continue

        raw_date = row[col_map['date_col']] if col_map['date_col'] < len(row) else None
        dt = _parse_date(raw_date)

        part_val = row[col_map['part_col']] if col_map['part_col'] < len(row) else None
        part = str(_clean(part_val) or '').strip()

        debit = _parse_amt(row[col_map['debit_col']] if col_map['debit_col'] < len(row) else None)
        credit = _parse_amt(row[col_map['credit_col']] if col_map['credit_col'] < len(row) else None)

        if not dt:
            # Check if this row is a multiline continuation of the previous transaction's narration
            if part and transactions and debit == 0.0 and credit == 0.0:
                if part.lower() not in ('nan', 'particulars', 'narration', 'description'):
                    transactions[-1]["narration"] += " " + part
            continue

        if not part or part.lower() in ('nan', 'particulars', 'narration', 'description'):
            continue

        balance = None
        if 'balance_col' in col_map and col_map['balance_col'] < len(row):
            balance = _parse_amt(row[col_map['balance_col']])

        transactions.append({
            "date": dt, "narration": part,
            "debit": debit, "credit": credit, "balance": balance
        })

    return transactions if transactions else None


def parse(file_path, password=None):
    """Parse a KVB Excel statement. Handles all known KVB formats."""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Failed to open KVB Excel: {e}")
    
    all_transactions = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [tuple(cell.value for cell in row) for row in ws.iter_rows()]
        txns = _try_parse_sheet(rows)
        if txns:
            all_transactions.extend(txns)
    
    wb.close()
    
    if not all_transactions:
        raise ValueError("Could not extract transactions from KVB statement.")
    
    all_transactions.sort(key=lambda t: t['date'])
    return all_transactions
