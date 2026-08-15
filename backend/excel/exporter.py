"""
Takes a standardized list of transaction dicts (as returned by any
bank parser) and:
  1. Classifies each row as Debit/Credit x plain/Charge/Interest
  2. Writes a complete, professional Excel workbook:
     - Summary (with Bank Name, Period, Opening/Closing Balance, Totals, Month breakdown)
     - All Transactions (S.No, Date, Particulars, Type, Debit, Credit, Balance, Reference, TOTAL row)
     - Inter-Company Transactions (S.No, Date, Particulars, Debit, Credit, Balance, Entity, Reference, Match Status, TOTAL row)
     - Monthly Sheets (Apr-2025, May-2025... with S.No, Date, Particulars, Type, Debit, Credit, Balance, TOTAL row)
"""

import re
from datetime import datetime
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CHARGE_RE = re.compile(r"\bCHG\b|CHARGE|\bFEE\b|\bTAX\b|\bGST\b|COMMISSION|FOLIO AMT|\bSC\b", re.IGNORECASE)
INTEREST_RE = re.compile(r"\bINT\b|INTEREST", re.IGNORECASE)


def classify_row(narration, debit, credit):
    side = "Debit" if (debit or 0) > 0 else ("Credit" if (credit or 0) > 0 else None)
    if side is None:
        return "Other"

    narration = narration or ""
    if INTEREST_RE.search(narration):
        category = "Interest"
    elif CHARGE_RE.search(narration):
        category = "Charge"
    else:
        category = None

    return f"{side} {category}" if category else side


def to_dataframe(transactions):
    df = pd.DataFrame(transactions)
    if df.empty:
        return df

    if "type" in df.columns:
        df["Type"] = df["type"]
    else:
        df["Type"] = df.apply(lambda r: classify_row(r.get("narration"), r.get("debit"), r.get("credit")), axis=1)

    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df["Month"] = df["date"].dt.month
        df["Year"] = df["date"].dt.year
        df["MonthLabel"] = df["date"].dt.strftime("%b-%Y")
        df = df.sort_values("date", kind="mergesort").reset_index(drop=True)
        df = df.rename(columns={"date": "Date"})

    if "narration" in df.columns:
        df = df.rename(columns={"narration": "Particulars"})
    if "debit" in df.columns:
        df = df.rename(columns={"debit": "Debit"})
    if "credit" in df.columns:
        df = df.rename(columns={"credit": "Credit"})
    if "balance" in df.columns:
        df = df.rename(columns={"balance": "Balance"})

    for col in ["Debit", "Credit", "Balance"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
            if col != "Balance":
                df[col] = df[col].fillna(0.0)

    return df



def _format_sheet(ws, is_table_sheet=True, header_row=1):
    """Apply styling, number formatting, freeze panes, autofit, and borders."""
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    total_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    total_border = Border(
        top=Side(style="thin", color="0F172A"),
        bottom=Side(style="double", color="0F172A"),
    )

    max_col = ws.max_column
    max_row = ws.max_row

    if is_table_sheet and header_row <= max_row:
        # Style table header
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left", vertical="center")

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"

    # Format data rows
    for r in range(header_row + 1 if is_table_sheet else 1, max_row + 1):
        # Check if TOTAL row
        first_cell_val = str(ws.cell(row=r, column=1).value or "").strip().upper()
        is_total_row = "TOTAL" in first_cell_val or str(ws.cell(row=r, column=2).value or "").strip().upper() == "TOTAL"

        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            header_val = str(ws.cell(row=header_row, column=c).value or "").strip().lower() if is_table_sheet else ""

            if is_total_row:
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border
            elif is_table_sheet:
                cell.border = thin_border

            # Number formatting
            if isinstance(cell.value, (int, float)):
                if any(k in header_val for k in ["debit", "credit", "balance", "charge", "interest", "amount", "total"]):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif "count" in header_val or "s.no" in header_val or header_val == "txns":
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            elif isinstance(cell.value, str):
                if any(k in header_val for k in ["date", "month", "s.no", "type", "status"]):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column widths
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                val_str = f"{val:,.2f}" if isinstance(val, float) else str(val)
                lines = val_str.split("\n")
                line_max = max(len(l) for l in lines) if lines else 0
                max_len = max(max_len, line_max)
        ws.column_dimensions[col_letter].width = max(min(max_len + 4, 60), 12)


def _month_summary_rows(df, month_labels):
    summary_rows = []
    for label in month_labels:
        month_df = df[df["MonthLabel"] == label]

        def sum_type(t, col):
            val = month_df.loc[month_df["Type"] == t, col].sum()
            return round(float(val), 2)

        summary_rows.append({
            "Month": label,
            "Credit": sum_type("Credit", "Credit"),
            "Debit": sum_type("Debit", "Debit"),
            "Credit Charge": sum_type("Credit Charge", "Credit"),
            "Debit Charge": sum_type("Debit Charge", "Debit"),
            "Credit Interest": sum_type("Credit Interest", "Credit"),
            "Debit Interest": sum_type("Debit Interest", "Debit"),
            "Transaction Count": len(month_df),
        })
    return summary_rows


def summarize(transactions):
    """Return JSON-serializable month-wise summary rows (for API responses)."""
    df = to_dataframe(transactions)
    if df.empty:
        return []

    month_labels = (
        df[["Year", "Month", "MonthLabel"]]
        .drop_duplicates()
        .sort_values(["Year", "Month"])["MonthLabel"]
        .tolist()
    )
    rows = _month_summary_rows(df, month_labels)

    totals = {"Month": "TOTAL (FY)"}
    for key in ["Credit", "Debit", "Credit Charge", "Debit Charge", "Credit Interest", "Debit Interest"]:
        totals[key] = round(sum(r[key] for r in rows), 2)
    totals["Transaction Count"] = sum(r["Transaction Count"] for r in rows)
    rows.append(totals)

    return rows


def summarize_combined(txns_A, txns_B=None):
    """
    Generate a combined summary table. Month-wise external credits and debits exclude inter-company transactions.
    """
    df_A = to_dataframe(txns_A)
    if txns_B:
        df_B = to_dataframe(txns_B)
        combined_df = pd.concat([df_A, df_B], ignore_index=True)
    else:
        combined_df = df_A

    if combined_df.empty:
        return []

    month_labels = (
        combined_df[["Year", "Month", "MonthLabel"]]
        .drop_duplicates()
        .sort_values(["Year", "Month"])["MonthLabel"]
        .tolist()
    )

    summary_rows = []
    for label in month_labels:
        month_df = combined_df[combined_df["MonthLabel"] == label]

        def sum_type(t, col):
            val = month_df.loc[month_df["Type"] == t, col].sum()
            return round(float(val), 2)

        raw_credit = sum_type("Credit", "Credit")
        raw_debit = sum_type("Debit", "Debit")

        # Inter-company logic
        ic_credit_sum = 0.0
        ic_debit_sum = 0.0
        if "inter_company" in month_df.columns:
            ic_credit_sum = round(float(month_df.loc[month_df["Type"] == "Credit", "inter_company"].sum()), 2)
            ic_debit_sum = round(float(month_df.loc[month_df["Type"] == "Debit", "inter_company"].sum()), 2)

        ic_total = max(ic_credit_sum, ic_debit_sum)
        ext_credit = max(0.0, round(raw_credit - ic_credit_sum, 2))
        ext_debit = max(0.0, round(raw_debit - ic_debit_sum, 2))

        summary_rows.append({
            "Month": label,
            "Credit": ext_credit,
            "Debit": ext_debit,
            "Credit Charge": sum_type("Credit Charge", "Credit"),
            "Debit Charge": sum_type("Debit Charge", "Debit"),
            "Credit Interest": sum_type("Credit Interest", "Credit"),
            "Debit Interest": sum_type("Debit Interest", "Debit"),
            "Inter-Company Transactions": ic_total,
            "Transaction Count": len(month_df),
        })

    totals = {"Month": "TOTAL (FY)"}
    for key in ["Credit", "Debit", "Credit Charge", "Debit Charge", "Credit Interest", "Debit Interest", "Inter-Company Transactions"]:
        totals[key] = round(sum(r[key] for r in summary_rows), 2)
    totals["Transaction Count"] = sum(r["Transaction Count"] for r in summary_rows)
    summary_rows.append(totals)

    return summary_rows


def verify_balance_chain(transactions):
    """
    Verify the running balance chain of the transactions.
    Supports both ascending and descending chronological orders.
    """
    if not transactions:
        return "no_balances", 0

    has_any_balance = any(t.get("balance") is not None for t in transactions)
    if not has_any_balance:
        return "no_balances", 0

    # 1. Forward direction test
    forward_mismatches = 0
    prev_bal = None
    for t in transactions:
        curr_bal = t.get("balance")
        debit = t.get("debit") or 0.0
        credit = t.get("credit") or 0.0
        if curr_bal is not None:
            if prev_bal is not None:
                expected = round(prev_bal - debit + credit, 2)
                actual = round(curr_bal, 2)
                if abs(expected - actual) > 0.02:
                    forward_mismatches += 1
            prev_bal = curr_bal

    # 2. Backward direction test (reverse chronological order)
    backward_mismatches = 0
    prev_bal = None
    for t in reversed(transactions):
        curr_bal = t.get("balance")
        debit = t.get("debit") or 0.0
        credit = t.get("credit") or 0.0
        if curr_bal is not None:
            if prev_bal is not None:
                expected = round(prev_bal - debit + credit, 2)
                actual = round(curr_bal, 2)
                if abs(expected - actual) > 0.02:
                    backward_mismatches += 1
            prev_bal = curr_bal

    mismatches = min(forward_mismatches, backward_mismatches)
    if mismatches > 0:
        return "mismatches", mismatches
    return "verified", 0


def _build_transaction_table_df(df):
    """Create formatted DataFrame with S.No and standard columns + TOTAL row."""
    if df.empty:
        return df

    out = df.copy()
    out["S.No"] = range(1, len(out) + 1)
    if "Date" in out.columns:
        out["Date"] = pd.to_datetime(out["Date"]).dt.strftime("%d-%m-%Y")

    if "Reference" not in out.columns:
        out["Reference"] = ""

    cols = ["S.No", "Date", "Particulars", "Type", "Debit", "Credit", "Balance", "Reference"]
    available_cols = [c for c in cols if c in out.columns]
    out = out[available_cols]

    total_debit = out["Debit"].sum() if "Debit" in out.columns else 0.0
    total_credit = out["Credit"].sum() if "Credit" in out.columns else 0.0

    total_row = {}
    for k in available_cols:
        if k in ["Debit", "Credit"]:
            total_row[k] = round(total_debit if k == "Debit" else total_credit, 2)
        elif k == "Date":
            total_row[k] = "TOTAL"
        else:
            total_row[k] = ""

    total_df = pd.DataFrame([total_row])
    # Ensure types match
    for num_col in ["Debit", "Credit"]:
        if num_col in total_df.columns:
            total_df[num_col] = pd.to_numeric(total_df[num_col])

    return pd.concat([out, total_df], ignore_index=True)



def export_to_excel(transactions, output_path, transactions_B=None, matched_pairs=None, unreconciled_logs=None, bank_name="Bank Statement"):
    """
    Generate the complete analysis workbook:
    1. Summary (Metadata cards + Monthly Summary Table)
    2. All Transactions (with S.No, Type, Debit, Credit, Balance, Reference, TOTAL row)
    3. Inter-Company Transactions (with S.No, Particulars, Debit, Credit, Entity, Status, TOTAL row)
    4. Monthly sheets (Apr-2025, May-2025... with S.No, TOTAL row)
    """
    df_A = to_dataframe(transactions)
    if df_A.empty:
        raise ValueError("No transactions to export.")

    if matched_pairs is None:
        matched_pairs = []
    if unreconciled_logs is None:
        unreconciled_logs = []

    is_dual = transactions_B is not None
    df_all = pd.concat([df_A, to_dataframe(transactions_B)], ignore_index=True) if is_dual else df_A

    # Calculate global metrics
    all_raw_txns = (transactions + transactions_B) if is_dual else transactions
    total_txns_count = len(df_all)
    total_debit = round(float(df_all["Debit"].sum()), 2)
    total_credit = round(float(df_all["Credit"].sum()), 2)

    # Opening & closing balances
    opening_bal = None
    for t in all_raw_txns:
        if t.get("balance") is not None:
            opening_bal = t["balance"]
            break

    closing_bal = None
    for t in reversed(all_raw_txns):
        if t.get("balance") is not None:
            closing_bal = t["balance"]
            break

    # Date range
    min_date = df_all["Date"].min() if "Date" in df_all.columns and not df_all.empty else None
    max_date = df_all["Date"].max() if "Date" in df_all.columns and not df_all.empty else None
    period_str = f"{pd.to_datetime(min_date).strftime('%d-%m-%Y')} to {pd.to_datetime(max_date).strftime('%d-%m-%Y')}" if min_date and max_date else "N/A"

    # Reconciliation status
    status_A, mismatches_A = verify_balance_chain(transactions)
    if is_dual:
        status_B, mismatches_B = verify_balance_chain(transactions_B)
        recon_status = "PASS" if (status_A == "verified" and status_B == "verified") else "REVIEW REQUIRED"
    else:
        recon_status = "PASS" if status_A == "verified" else "REVIEW REQUIRED"

    # Inter-company transactions subset
    ic_df = df_all[df_all.get("is_inter_company", False) | (df_all.get("inter_company", 0.0) > 0)].copy()
    ic_count = len(ic_df)
    ic_debit = round(float(ic_df["Debit"].sum()), 2) if not ic_df.empty else 0.0
    ic_credit = round(float(ic_df["Credit"].sum()), 2) if not ic_df.empty else 0.0

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # =========================================================================
        # 1. SUMMARY SHEET
        # =========================================================================
        summary_rows = summarize_combined(transactions, transactions_B) if is_dual else summarize(transactions)
        summary_df = pd.DataFrame(summary_rows)

        # Write metadata card section at top of Summary sheet
        metadata_records = [
            ("BANK STATEMENT ANALYSIS REPORT", ""),
            ("Bank Name", bank_name),
            ("Statement Period", period_str),
            ("Total Transactions", total_txns_count),
            ("Total Debit", total_debit),
            ("Total Credit", total_credit),
            ("Opening Balance", opening_bal if opening_bal is not None else "N/A"),
            ("Closing Balance", closing_bal if closing_bal is not None else "N/A"),
            ("Reconciliation Status", recon_status),
            ("Inter-Company Transactions Count", ic_count),
            ("Inter-Company Total Debit", ic_debit),
            ("Inter-Company Total Credit", ic_credit),
            ("", ""),  # spacer
            ("MONTH-WISE BREAKDOWN SUMMARY", ""),
        ]

        meta_df = pd.DataFrame(metadata_records, columns=["Metric", "Value"])
        meta_df.to_excel(writer, sheet_name="Summary", startrow=0, index=False)

        # Write monthly breakdown table below metadata
        summary_start_row = len(metadata_records) + 2
        summary_df.to_excel(writer, sheet_name="Summary", startrow=summary_start_row, index=False)

        ws_sum = writer.sheets["Summary"]
        # Format metadata section
        title_font = Font(name="Calibri", size=14, bold=True, color="0F5C4D")
        sub_font = Font(name="Calibri", size=12, bold=True, color="1E293B")
        meta_label_font = Font(name="Calibri", size=11, bold=True, color="475569")
        meta_val_font = Font(name="Calibri", size=11, bold=True, color="0F172A")

        ws_sum.cell(row=2, column=1).font = title_font
        ws_sum.cell(row=len(metadata_records) + 2, column=1).font = sub_font

        for r in range(3, len(metadata_records) + 1):
            ws_sum.cell(row=r, column=1).font = meta_label_font
            ws_sum.cell(row=r, column=2).font = meta_val_font
            val = ws_sum.cell(row=r, column=2).value
            if isinstance(val, float):
                ws_sum.cell(row=r, column=2).number_format = "#,##0.00"

        _format_sheet(ws_sum, is_table_sheet=False)

        # Style the monthly summary table headers and totals
        table_header_row = summary_start_row + 1
        for col_idx in range(1, len(summary_df.columns) + 1):
            cell = ws_sum.cell(row=table_header_row, column=col_idx)
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style totals row and format table amounts
        last_table_row = table_header_row + len(summary_df)
        for r in range(table_header_row + 1, last_table_row + 1):
            is_total = r == last_table_row
            for c in range(1, len(summary_df.columns) + 1):
                cell = ws_sum.cell(row=r, column=c)
                if is_total:
                    cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                    cell.font = Font(name="Calibri", size=11, bold=True, color="0F172A")
                    cell.border = Border(top=Side(style="thin", color="0F172A"), bottom=Side(style="double", color="0F172A"))
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"

        # Write matched pairs / unreconciled notes at bottom of summary
        note_row = last_table_row + 3
        if matched_pairs:
            ws_sum.cell(row=note_row, column=1, value="Matched Inter-Company Pairs:").font = sub_font
            for p in matched_pairs:
                note_row += 1
                ws_sum.cell(row=note_row, column=1, value=str(p))

        if unreconciled_logs:
            note_row += 2
            ws_sum.cell(row=note_row, column=1, value="Unreconciled Transactions:").font = sub_font
            for u in unreconciled_logs:
                note_row += 1
                ws_sum.cell(row=note_row, column=1, value=str(u))

        # =========================================================================
        # 2. ALL TRANSACTIONS SHEET
        # =========================================================================
        all_txns_table_df = _build_transaction_table_df(df_all)
        all_txns_table_df.to_excel(writer, sheet_name="All Transactions", index=False)
        _format_sheet(writer.sheets["All Transactions"], is_table_sheet=True)

        # =========================================================================
        # 3. INTER-COMPANY TRANSACTIONS SHEET
        # =========================================================================
        if not ic_df.empty:
            ic_table_df = _build_transaction_table_df(ic_df)
        else:
            # Empty placeholder table with columns
            ic_table_df = pd.DataFrame(columns=["S.No", "Date", "Particulars", "Type", "Debit", "Credit", "Balance", "Reference"])
            empty_row = pd.DataFrame([{"S.No": "", "Date": "No inter-company transactions detected", "Particulars": "", "Type": "", "Debit": 0.0, "Credit": 0.0, "Balance": None, "Reference": ""}])
            ic_table_df = pd.concat([ic_table_df, empty_row], ignore_index=True)

        ic_table_df.to_excel(writer, sheet_name="Inter-Company Transactions", index=False)
        _format_sheet(writer.sheets["Inter-Company Transactions"], is_table_sheet=True)

        # =========================================================================
        # 4. MONTHLY TRANSACTION SHEETS (Apr-2025, May-2025, etc.)
        # =========================================================================
        if "MonthLabel" in df_all.columns:
            month_labels = (
                df_all[["Year", "Month", "MonthLabel"]]
                .dropna()
                .drop_duplicates()
                .sort_values(["Year", "Month"])["MonthLabel"]
                .tolist()
            )
            for label in month_labels:
                m_df = df_all[df_all["MonthLabel"] == label].copy()
                if not m_df.empty:
                    m_table_df = _build_transaction_table_df(m_df)
                    sheet_name = label[:31]
                    m_table_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    _format_sheet(writer.sheets[sheet_name], is_table_sheet=True)

        # Dual statement individual sheets if present
        if is_dual:
            sheet_A_df = _build_transaction_table_df(df_A)
            sheet_A_df.to_excel(writer, sheet_name="Statement A", index=False)
            _format_sheet(writer.sheets["Statement A"], is_table_sheet=True)

            sheet_B_df = _build_transaction_table_df(to_dataframe(transactions_B))
            sheet_B_df.to_excel(writer, sheet_name="Statement B", index=False)
            _format_sheet(writer.sheets["Statement B"], is_table_sheet=True)

        # =========================================================================
        # REORDER SHEETS: Summary -> All Transactions -> Inter-Company -> Monthly
        # =========================================================================
        desired_order = ["Summary", "All Transactions", "Inter-Company Transactions"]
        current_sheets = writer.book.sheetnames
        for idx, sname in enumerate(desired_order):
            if sname in current_sheets:
                cur_idx = writer.book.sheetnames.index(sname)
                offset = idx - cur_idx
                if offset != 0:
                    writer.book.move_sheet(sname, offset=offset)

    return output_path
