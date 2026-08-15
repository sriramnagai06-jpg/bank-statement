"""
Comprehensive Test Suite for Bank Statement Analyzer
Tests:
  - Classification and Debit/Credit integrity
  - Multiline narration parsing
  - Inter-company transaction detection and retention
  - Balance reconciliation verification
  - Multi-sheet Excel workbook export (Summary, All Transactions, Inter-Company, Monthly)
  - API output consistency (One Source of Truth)
"""

import os
import tempfile
import unittest
from datetime import datetime
import openpyxl
import pandas as pd

import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

from excel.exporter import (
    classify_row,
    export_to_excel,
    summarize,
    summarize_combined,
    to_dataframe,
    verify_balance_chain,
)
from parser.copypaste import parse_text
from app import classify_transactions


class TestBankStatementAnalyzer(unittest.TestCase):

    def setUp(self):
        self.sample_txns = [
            {
                "date": datetime(2025, 4, 1),
                "narration": "Opening Balance B/F",
                "debit": 0.0,
                "credit": 0.0,
                "balance": 10000.0,
            },
            {
                "date": datetime(2025, 4, 2),
                "narration": "UPI/DR/12345/Rent Payment to Landlord",
                "debit": 3000.0,
                "credit": 0.0,
                "balance": 7000.0,
            },
            {
                "date": datetime(2025, 4, 10),
                "narration": "NEFT/CR/67890/Client Consulting Fees",
                "debit": 0.0,
                "credit": 15000.0,
                "balance": 22000.0,
            },
            {
                "date": datetime(2025, 4, 15),
                "narration": "Transfer to Lakshmi Traders Inter-Company",
                "debit": 5000.0,
                "credit": 0.0,
                "balance": 17000.0,
            },
            {
                "date": datetime(2025, 4, 20),
                "narration": "SMS CHG AND GST TAX",
                "debit": 17.70,
                "credit": 0.0,
                "balance": 16982.30,
            },
            {
                "date": datetime(2025, 5, 5),
                "narration": "SAVINGS BANK INT RECEIVED",
                "debit": 0.0,
                "credit": 250.0,
                "balance": 17232.30,
            },
        ]

    def test_classify_row(self):
        self.assertEqual(classify_row("UPI Rent", 1000.0, 0.0), "Debit")
        self.assertEqual(classify_row("Salary Deposit", 0.0, 5000.0), "Credit")
        self.assertEqual(classify_row("SMS ALERT CHG", 15.0, 0.0), "Debit Charge")
        self.assertEqual(classify_row("SAVINGS INT", 0.0, 300.0), "Credit Interest")

    def test_classify_transactions_preserves_all_rows(self):
        txns = [dict(t) for t in self.sample_txns]
        initial_count = len(txns)
        classified = classify_transactions(txns)
        
        # Verify NO transactions were deleted
        self.assertEqual(len(classified), initial_count)
        
        # Verify inter-company is flagged
        ic_txns = [t for t in classified if t.get("is_inter_company")]
        self.assertEqual(len(ic_txns), 1)
        self.assertEqual(ic_txns[0]["debit"], 5000.0)

    def test_balance_chain_verification(self):
        status, mismatches = verify_balance_chain(self.sample_txns)
        self.assertEqual(status, "verified")
        self.assertEqual(mismatches, 0)

        # Break a balance deliberately
        bad_txns = [dict(t) for t in self.sample_txns]
        bad_txns[2]["balance"] = 99999.0
        bad_status, bad_mismatches = verify_balance_chain(bad_txns)
        self.assertEqual(bad_status, "mismatches")
        self.assertGreater(bad_mismatches, 0)

    def test_multiline_copypaste_parsing(self):
        pasted = (
            "04-04-2025 UPI/DR/546084542514/RAVI 4,300.00 6,833.34\n"
            "N/ESFB/**NAGAI@OKAXIS/UPI/\n"
            "ICIC39BD5E2150F41DDAB04C\n"
            "05-04-2025 UPI/CR/546076693044/PRABU 4,300.00 Cr 11,133.34\n"
        )
        parsed = parse_text(pasted)
        self.assertEqual(len(parsed), 2)
        # Verify continuation lines merged into narration of first txn
        self.assertIn("NAGAI@OKAXIS", parsed[0]["narration"])
        self.assertEqual(parsed[0]["debit"], 4300.0)
        self.assertEqual(parsed[1]["credit"], 4300.0)

    def test_excel_export_complete_structure(self):
        txns = [dict(t) for t in self.sample_txns]
        classify_transactions(txns)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            export_to_excel(txns, tmp_path, bank_name="State Bank of India")

            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            sheet_names = wb.sheetnames

            # 1. Verify required sheets exist and are in the correct order
            self.assertEqual(sheet_names[0], "Summary")
            self.assertEqual(sheet_names[1], "All Transactions")
            self.assertEqual(sheet_names[2], "Inter-Company Transactions")
            self.assertIn("Apr-2025", sheet_names)
            self.assertIn("May-2025", sheet_names)

            # 2. Verify All Transactions sheet
            ws_all = wb["All Transactions"]
            headers = [ws_all.cell(row=1, column=c).value for c in range(1, ws_all.max_column + 1)]
            self.assertIn("S.No", headers)
            self.assertIn("Date", headers)
            self.assertIn("Particulars", headers)
            self.assertIn("Debit", headers)
            self.assertIn("Credit", headers)
            self.assertIn("Balance", headers)

            # Check bottom TOTAL row
            last_row = ws_all.max_row
            date_col_idx = headers.index("Date") + 1
            self.assertEqual(ws_all.cell(row=last_row, column=date_col_idx).value, "TOTAL")

            debit_col_idx = headers.index("Debit") + 1
            credit_col_idx = headers.index("Credit") + 1
            total_debit_excel = ws_all.cell(row=last_row, column=debit_col_idx).value
            total_credit_excel = ws_all.cell(row=last_row, column=credit_col_idx).value

            expected_debit = sum(t["debit"] for t in txns)
            expected_credit = sum(t["credit"] for t in txns)
            self.assertAlmostEqual(total_debit_excel, expected_debit, places=2)
            self.assertAlmostEqual(total_credit_excel, expected_credit, places=2)

            # 3. Verify Inter-Company Transactions sheet
            ws_ic = wb["Inter-Company Transactions"]
            ic_headers = [ws_ic.cell(row=1, column=c).value for c in range(1, ws_ic.max_column + 1)]
            self.assertIn("S.No", ic_headers)
            self.assertIn("Debit", ic_headers)
            # 1 IC txn + 1 header + 1 total row = 3 rows
            self.assertEqual(ws_ic.max_row, 3)

            wb.close()
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)


if __name__ == "__main__":
    unittest.main()
